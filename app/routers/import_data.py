"""スタッフ・セッションのインポート（Excel/CSV）とテンプレートのダウンロード。"""

import csv
import io
import re
import traceback
from datetime import datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from ..config import now as app_now
from ..database import get_db
from ..labels import build_label_maps
from ..models import Room, SessionGroup, Staff, StaffAvailability
from ..models import Session as SessionModel
from .staffs import _check_availability_range, _event_range

from .backup import _read_limited

# 取り込むファイルの上限
MAX_IMPORT_BYTES = 50 * 1024 * 1024

router = APIRouter(prefix="/api/export", tags=["export"])

# エクスポートのシートと同じ列構成（テンプレートと必須列の単一ソース）
STAFF_HEADERS = ["ID", "名前", "Slack名", "担当", "経験回数", "英語", "最大稼働時間", "活動可能時間", "担当セッション数"]
STAFF_REQUIRED = ["名前"]
SESSION_HEADERS = ["ID", "写真", "タイトル", "登壇者", "ふりがな", "所属", "肩書き", "開始", "終了",
                   "部屋", "カテゴリ", "グループ", "必要人数", "英語", "説明", "備考"]
SESSION_REQUIRED = ["タイトル", "登壇者", "開始", "終了", "部屋"]

STAFF_SAMPLE = ["", "山田 太郎", "yamada", "セッション", 1, "○", "8h", "9/1 09:00-18:00 / 9/2 09:00-12:00", ""]
SESSION_SAMPLE = ["", "", "クラウドネイティブ入門", "山田 太郎", "やまだ たろう", "Example社", "エンジニア",
                  "2026/09/01 10:00", "2026/09/01 10:45", "ホールA", "一般", "セッション", 2, "○", "", ""]


# ---------------------------------------------------------------------------
# ファイル読み取り
# ---------------------------------------------------------------------------

def _decode_csv(raw: bytes) -> str:
    try:
        return raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            return raw.decode("cp932")
        except UnicodeDecodeError:
            raise HTTPException(status_code=400, detail="UTF-8のCSVまたはExcel(xlsx)ファイルを指定してください")


def _load_tables(raw: bytes, sheet_name: str) -> list[tuple[str | None, list[list]]]:
    """(シート名, 行リスト) のリストを返す。xlsxはシート名の前方一致、CSVは1テーブル。"""
    if raw[:4] == b"PK\x03\x04":
        try:
            wb = load_workbook(io.BytesIO(raw), data_only=True)
        except Exception:
            raise HTTPException(status_code=400, detail="Excelファイルを読み込めませんでした")
        names = [n for n in wb.sheetnames if n == sheet_name or n.startswith(sheet_name + " ")]
        if not names:
            names = [wb.sheetnames[0]]
        return [(n, [list(r) for r in wb[n].iter_rows(values_only=True)]) for n in names]
    text = _decode_csv(raw)
    rows = [row for row in csv.reader(io.StringIO(text))]
    return [(None, rows)]


def _header_map(header_row: list, required: list[str]) -> dict[str, int]:
    cols = {}
    for i, v in enumerate(header_row):
        name = str(v).strip() if v is not None else ""
        if name and name not in cols:
            cols[name] = i
    missing = [h for h in required if h not in cols]
    if missing:
        raise HTTPException(status_code=400, detail=f"必須の列がありません: {'、'.join(missing)}")
    return cols


def _cell(row: list, cols: dict, name: str):
    i = cols.get(name)
    if i is None or i >= len(row):
        return None
    return row[i]


def _text(row: list, cols: dict, name: str) -> str:
    v = _cell(row, cols, name)
    if v is None:
        return ""
    return str(v).strip()


# ---------------------------------------------------------------------------
# 値のパース
# ---------------------------------------------------------------------------

_CIRCLE_VALUES = {"○", "◯", "o", "true", "1", "yes", "はい"}


def _parse_circle(v) -> int:
    if v is None:
        return 0
    return 1 if str(v).strip().lower() in _CIRCLE_VALUES else 0


def _parse_hours(v) -> int:
    if v is None or str(v).strip() == "":
        return 8
    m = re.match(r"^\s*(\d+)(?:\.0+)?\s*h?\s*$", str(v).strip(), re.IGNORECASE)
    if not m:
        raise ValueError(f"最大稼働時間の形式が不正です: '{v}'")
    return int(m.group(1))


def _parse_int(v, default: int) -> int:
    if v is None or str(v).strip() == "":
        return default
    try:
        return int(float(str(v).strip()))
    except ValueError:
        raise ValueError(f"数値の形式が不正です: '{v}'")


_DT_FORMATS = ("%Y/%m/%d %H:%M", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M")


def _parse_dt_cell(v, label: str) -> datetime:
    if isinstance(v, datetime):
        return v
    text = str(v).strip() if v is not None else ""
    if not text:
        raise ValueError(f"{label}が入力されていません")
    for fmt in _DT_FORMATS:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(text)
    except ValueError:
        raise ValueError(f"{label}の形式が不正です: '{text}'")


_AVAIL_RE = re.compile(r"^(\d{1,2})/(\d{1,2})\s+(\d{1,2}):(\d{2})-(\d{1,2}):(\d{2})$")


def _parse_avails(db: Session, text: str) -> list[tuple[datetime, datetime]]:
    """エクスポート形式 "MM/DD HH:MM-HH:MM / ..." を解釈する。年は開催年で補完する。"""
    if not text:
        return []
    rng = _event_range(db)
    year = rng[0].year if rng else app_now().year
    out = []
    for part in text.split(" / "):
        part = part.strip()
        if not part:
            continue
        m = _AVAIL_RE.match(part)
        if not m:
            raise ValueError(f"活動可能時間の形式が不正です: '{part}'")
        mo, d, h1, mi1, h2, mi2 = (int(x) for x in m.groups())
        try:
            start = datetime(year, mo, d, h1, mi1)
            end = datetime(year, mo, d, h2, mi2)
        except ValueError:
            raise ValueError(f"活動可能時間の日付が不正です: '{part}'")
        if start >= end:
            raise ValueError(f"活動可能時間の開始が終了以降です: '{part}'")
        try:
            _check_availability_range(db, start, end)
        except HTTPException as e:
            raise ValueError(e.detail)
        out.append((start, end))
    return out


def _invert(labels: dict) -> dict:
    """label -> key の逆写像（後勝ち = ユーザー定義優先）"""
    return {v: k for k, v in labels.items()}


def _resolve(token: str, label_to_key: dict, keys: set) -> str | None:
    if token in label_to_key:
        return label_to_key[token]
    if token in keys:
        return token
    return None


# ---------------------------------------------------------------------------
# インポート本体
# ---------------------------------------------------------------------------

def _entry(sheet: str | None, row: int, name: str, reason: str) -> dict:
    e = {"row": row, "name": name, "reason": reason}
    if sheet:
        e["sheet"] = sheet
    return e


@router.post("/import-staffs")
async def import_staffs(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """スタッフをExcel/CSVから一括登録する。同名はスキップ。"""
    raw = await _read_limited(file, MAX_IMPORT_BYTES, "ファイル")
    tables = _load_tables(raw, "スタッフ管理")

    _, role_labels, _ = build_label_maps(db)
    role_to_key = _invert(role_labels)
    role_keys = set(role_labels.keys())

    seen_names = {s.name for s in db.query(Staff).all()}
    created = 0
    skipped: list[dict] = []
    errors: list[dict] = []
    try:
        for sheet, rows in tables:
            if not rows:
                continue
            cols = _header_map(rows[0], STAFF_REQUIRED)
            for idx, row in enumerate(rows[1:], start=2):
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                name = _text(row, cols, "名前")
                if not name:
                    errors.append(_entry(sheet, idx, "", "名前が入力されていません"))
                    continue
                if name in seen_names:
                    skipped.append(_entry(sheet, idx, name, "同名のスタッフが登録済み"))
                    continue
                try:
                    roles = []
                    role_text = _text(row, cols, "担当")
                    if role_text and role_text != "なし":
                        for token in re.split(r"[、,]", role_text):
                            token = token.strip()
                            if not token:
                                continue
                            key = _resolve(token, role_to_key, role_keys)
                            if key is None:
                                raise ValueError(f"担当が見つかりません: '{token}'")
                            if key not in roles:
                                roles.append(key)
                    max_hours = _parse_hours(_cell(row, cols, "最大稼働時間"))
                    experience = _parse_int(_cell(row, cols, "経験回数"), 0)
                    english = _parse_circle(_cell(row, cols, "英語"))
                    avails = _parse_avails(db, _text(row, cols, "活動可能時間"))
                except ValueError as e:
                    errors.append(_entry(sheet, idx, name, str(e)))
                    continue
                staff = Staff(
                    name=name,
                    slack_name=_text(row, cols, "Slack名"),
                    role=",".join(roles),
                    english_ok=english,
                    max_hours=max_hours,
                    experience_count=experience,
                )
                db.add(staff)
                db.flush()
                for start, end in avails:
                    db.add(StaffAvailability(staff_id=staff.id, start_time=start, end_time=end))
                seen_names.add(name)
                created += 1
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"detail": "インポート中にエラーが発生しました"})
    return {"status": "ok", "created": created, "skipped": skipped, "errors": errors}


@router.post("/import-sessions")
async def import_sessions(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """セッションをExcel/CSVから一括登録する。同タイトル・同開始時刻はスキップ。
    未登録の部屋・グループは自動作成する。"""
    raw = await _read_limited(file, MAX_IMPORT_BYTES, "ファイル")
    tables = _load_tables(raw, "セッション管理")

    cat_labels, _, _ = build_label_maps(db)
    cat_to_key = _invert(cat_labels)
    cat_keys = set(cat_labels.keys())

    room_by_name = {r.name: r for r in db.query(Room).all()}
    group_by_label = {g.label: g for g in db.query(SessionGroup).all()}
    max_group_order = max((g.order or 0 for g in group_by_label.values()), default=0)
    seen_sessions = {(s.title, s.start_time) for s in db.query(SessionModel).all()}

    created = 0
    created_rooms: list[str] = []
    created_groups: list[str] = []
    skipped: list[dict] = []
    errors: list[dict] = []
    try:
        for sheet, rows in tables:
            if not rows:
                continue
            cols = _header_map(rows[0], SESSION_REQUIRED)
            for idx, row in enumerate(rows[1:], start=2):
                if all(v is None or str(v).strip() == "" for v in row):
                    continue
                title = _text(row, cols, "タイトル")
                if not title:
                    errors.append(_entry(sheet, idx, "", "タイトルが入力されていません"))
                    continue
                try:
                    start = _parse_dt_cell(_cell(row, cols, "開始"), "開始")
                    end = _parse_dt_cell(_cell(row, cols, "終了"), "終了")
                    if start >= end:
                        raise ValueError("開始時刻は終了時刻より前にしてください")
                    speaker = _text(row, cols, "登壇者")
                    if not speaker:
                        raise ValueError("登壇者が入力されていません")
                    room_name = _text(row, cols, "部屋")
                    if not room_name:
                        raise ValueError("部屋が入力されていません")
                    cat_text = _text(row, cols, "カテゴリ")
                    if cat_text:
                        category = _resolve(cat_text, cat_to_key, cat_keys)
                        if category is None:
                            raise ValueError(f"カテゴリが見つかりません: '{cat_text}'")
                        if category == "overall":
                            raise ValueError("全体カテゴリはインポート対象外です")
                    else:
                        category = "general"
                    required = _parse_int(_cell(row, cols, "必要人数"), 1)
                    english = _parse_circle(_cell(row, cols, "英語"))
                except ValueError as e:
                    errors.append(_entry(sheet, idx, title, str(e)))
                    continue
                if (title, start) in seen_sessions:
                    skipped.append(_entry(sheet, idx, title, "同タイトル・同開始時刻のセッションが登録済み"))
                    continue
                # 検証がすべて通ってから部屋・グループを自動作成する
                room = room_by_name.get(room_name)
                if room is None:
                    room = Room(name=room_name, capacity=0, floor=1)
                    db.add(room)
                    db.flush()
                    room_by_name[room_name] = room
                    created_rooms.append(room_name)
                group_id = None
                group_label = _text(row, cols, "グループ")
                if group_label:
                    group = group_by_label.get(group_label)
                    if group is None:
                        max_group_order += 1
                        group = SessionGroup(label=group_label, date="", color="#1a73e8", order=max_group_order)
                        db.add(group)
                        db.flush()
                        group_by_label[group_label] = group
                        created_groups.append(group_label)
                    group_id = group.id
                db.add(SessionModel(
                    title=title,
                    speaker=speaker,
                    speaker_kana=_text(row, cols, "ふりがな"),
                    speaker_org=_text(row, cols, "所属"),
                    speaker_title=_text(row, cols, "肩書き"),
                    start_time=start,
                    end_time=end,
                    room_id=room.id,
                    category=category,
                    required_staff=required,
                    english_required=english,
                    description=_text(row, cols, "説明"),
                    notes=_text(row, cols, "備考"),
                    group_id=group_id,
                ))
                seen_sessions.add((title, start))
                created += 1
        db.commit()
    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        traceback.print_exc()
        return JSONResponse(status_code=500, content={"detail": "インポート中にエラーが発生しました"})
    return {
        "status": "ok", "created": created, "skipped": skipped, "errors": errors,
        "created_rooms": created_rooms, "created_groups": created_groups,
    }


# ---------------------------------------------------------------------------
# テンプレート
# ---------------------------------------------------------------------------

_TEMPLATES = {
    "staffs": ("スタッフ管理", STAFF_HEADERS, STAFF_SAMPLE, "staff_import_template"),
    "sessions": ("セッション管理", SESSION_HEADERS, SESSION_SAMPLE, "session_import_template"),
}


@router.get("/template/{kind}")
def download_template(kind: str, fmt: str = "xlsx"):
    """インポート用テンプレート（ヘッダ + 記入例1行）をダウンロードする"""
    if kind not in _TEMPLATES:
        raise HTTPException(status_code=404, detail="Not found")
    sheet_name, headers, sample, basename = _TEMPLATES[kind]
    if fmt == "csv":
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(headers)
        writer.writerow(sample)
        data = buf.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(data), media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={basename}.csv"},
        )
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    ws.append(sample)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={basename}.xlsx"},
    )
