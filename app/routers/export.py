import io
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from sqlalchemy.orm import Session, joinedload

from ..config import UPLOAD_DIR, now as app_now
from ..database import get_db
from ..models import (
    Session as SessionModel, Staff, Assignment, Room,
    LTTalk, Category, SessionGroup,
)
router = APIRouter(prefix="/api/export", tags=["export"])

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="388E3C", end_color="388E3C", fill_type="solid")
HEADER_FILL_PURPLE = PatternFill(start_color="7B1FA2", end_color="7B1FA2", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
HEADER_FILL_TEAL = PatternFill(start_color="00695C", end_color="00695C", fill_type="solid")
HEADER_FILL_BROWN = PatternFill(start_color="5D4037", end_color="5D4037", fill_type="solid")
HEADER_FILL_INDIGO = PatternFill(start_color="283593", end_color="283593", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")

SESSION_CATS = ("general", "tech", "workshop", "keynote", "lt", "panel")
ROLE_LABELS_BASE = {"session": "セッション"}
CAT_LABELS_BASE = {
    "general": "一般", "tech": "技術", "workshop": "ワークショップ",
    "keynote": "基調講演", "lt": "LT", "panel": "パネルディスカッション", "overall": "全体",
}

PHOTO_PX = 48  # Excel内の写真サイズ (px)
ROW_HEIGHT_WITH_PHOTO = 45  # 写真付き行の高さ (pt)


def _fmt(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%H:%M")


def _fmt_full(dt: datetime | None) -> str:
    if not dt:
        return ""
    return dt.strftime("%Y/%m/%d %H:%M")


def _apply_header(ws, row, fill=HEADER_FILL):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _apply_border(ws, row):
    for cell in ws[row]:
        cell.border = THIN_BORDER
        cell.alignment = WRAP


def _auto_width(ws, max_rows=100):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for i, cell in enumerate(col):
            if i >= max_rows:
                break
            try:
                val = str(cell.value or "")
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)


def _add_photo(ws, photo_path: str, col: int, row: int):
    """ワークシートのセルに元画像をそのまま埋め込む（セル内に収まるようリサイズ表示）"""
    file_path = Path("." + photo_path)
    if not file_path.exists():
        return
    try:
        img = PILImage.open(file_path)
        img.verify()  # 画像ファイルとして有効か確認
    except Exception:
        return
    xl_img = XlImage(str(file_path))
    xl_img.width = PHOTO_PX
    xl_img.height = PHOTO_PX
    cell_ref = f"{get_column_letter(col)}{row}"
    ws.add_image(xl_img, cell_ref)
    ws.row_dimensions[row].height = ROW_HEIGHT_WITH_PHOTO


@router.get("/excel")
def export_excel(db: Session = Depends(get_db)):
    """タブ構成に合わせてExcelファイルをエクスポート"""

    # データ取得
    rooms = db.query(Room).order_by(Room.id).all()
    sessions = (
        db.query(SessionModel)
        .options(
            joinedload(SessionModel.room),
            joinedload(SessionModel.lt_talks),
            joinedload(SessionModel.assignments).joinedload(Assignment.staff).joinedload(Staff.skills),
        )
        .order_by(SessionModel.start_time)
        .all()
    )
    staffs = (
        db.query(Staff)
        .options(
            joinedload(Staff.availabilities),
            joinedload(Staff.preferred_sessions),
            joinedload(Staff.assignments).joinedload(Assignment.session).joinedload(SessionModel.room),
        )
        .order_by(Staff.id)
        .all()
    )

    # 動的カテゴリ取得
    db_categories = db.query(Category).order_by(Category.order, Category.id).all()
    db_session_groups = db.query(SessionGroup).order_by(SessionGroup.order, SessionGroup.id).all()
    dynamic_cat_keys = [c.key for c in db_categories]
    CAT_LABELS = {**CAT_LABELS_BASE, **{c.key: c.label for c in db_categories}}
    ROLE_LABELS = {**ROLE_LABELS_BASE, **{c.key: c.label for c in db_categories}}

    def _cat_header_fill(color_hex: str) -> PatternFill:
        """カテゴリの色からヘッダー用PatternFillを生成"""
        c = color_hex.lstrip("#").upper()
        return PatternFill(start_color=c, end_color=c, fill_type="solid")

    cat_fill_map = {c.key: _cat_header_fill(c.color) for c in db_categories}

    wb = Workbook()

    # ============================================================
    # Sheet 1: 部屋管理
    # ============================================================
    ws1 = wb.active
    ws1.title = "部屋管理"
    ws1.append(["ID", "部屋名", "定員", "階"])
    _apply_header(ws1, 1, HEADER_FILL_BROWN)

    for r in rooms:
        ws1.append([r.id, r.name, r.capacity, r.floor])
        _apply_border(ws1, ws1.max_row)

    _auto_width(ws1)

    # ============================================================
    # Sheet 2: セッション管理 (動的カテゴリ除外, 写真付き)
    # ============================================================
    group_label_map = {g.id: g.label for g in db_session_groups}
    ws2 = wb.create_sheet("セッション管理")
    ws2.append(["ID", "写真", "タイトル", "登壇者", "ふりがな", "所属", "肩書き", "開始", "終了", "部屋", "カテゴリ", "グループ", "必要人数", "英語", "説明", "備考"])
    _apply_header(ws2, 1)
    # 写真列の幅を確保
    ws2.column_dimensions["B"].width = 9

    for s in sessions:
        if s.category not in SESSION_CATS:
            continue
        if s.category in ("lt", "panel") and s.lt_talks:
            speakers = "\n".join(
                f"{t.speaker}（{t.title}）" + (f" {t.start_time}〜{t.end_time}" if t.start_time else "")
                for t in s.lt_talks
            )
        else:
            speakers = s.speaker
        ws2.append([
            s.id,
            "",  # 写真列（画像で上書き）
            s.title,
            speakers,
            s.speaker_kana,
            s.speaker_org,
            s.speaker_title,
            _fmt_full(s.start_time),
            _fmt_full(s.end_time),
            s.room.name if s.room else "",
            CAT_LABELS.get(s.category, s.category),
            group_label_map.get(s.group_id, ""),
            s.required_staff,
            "○" if s.english_required else "",
            s.description,
            s.notes,
        ])
        _apply_border(ws2, ws2.max_row)
        if s.speaker_photo:
            _add_photo(ws2, s.speaker_photo, col=2, row=ws2.max_row)

    _auto_width(ws2)
    ws2.column_dimensions["B"].width = 9  # auto_widthで上書きされるので再設定

    # ============================================================
    # Sheet 3: スタッフ管理
    # ============================================================
    ws3 = wb.create_sheet("スタッフ管理")
    ws3.append(["ID", "名前", "Slack名", "担当", "経験回数", "英語", "最大稼働時間", "活動可能時間", "担当セッション数"])
    _apply_header(ws3, 1, HEADER_FILL_ORANGE)

    for st in staffs:
        avails = " / ".join(
            f"{_fmt(a.start_time)}-{_fmt(a.end_time)}" for a in st.availabilities
        )
        ws3.append([
            st.id,
            st.name,
            st.slack_name,
            ROLE_LABELS.get(st.role, st.role),
            st.experience_count,
            "○" if st.english_ok else "",
            f"{st.max_hours}h",
            avails,
            len(st.assignments),
        ])
        _apply_border(ws3, ws3.max_row)

    _auto_width(ws3)

    # ============================================================
    # Sheet 4: セッション配置 (写真付き)
    # ============================================================
    ws4 = wb.create_sheet("セッション配置")
    ws4.append(["時間", "部屋", "写真", "タイトル", "登壇者", "カテゴリ", "必要人数", "配置人数", "英語", "担当スタッフ", "備考"])
    _apply_header(ws4, 1, HEADER_FILL_TEAL)
    ws4.column_dimensions["C"].width = 9

    for s in sessions:
        if s.category not in SESSION_CATS:
            continue
        staff_names = ", ".join(a.staff.name for a in s.assignments)
        assigned_count = len(s.assignments)
        status = "○" if assigned_count >= s.required_staff else f"不足({assigned_count}/{s.required_staff})"
        ws4.append([
            f"{_fmt(s.start_time)}-{_fmt(s.end_time)}",
            s.room.name if s.room else "",
            "",  # 写真列
            s.title,
            s.speaker,
            CAT_LABELS.get(s.category, s.category),
            s.required_staff,
            status,
            "○" if s.english_required else "",
            staff_names,
            s.notes,
        ])
        _apply_border(ws4, ws4.max_row)
        if s.speaker_photo:
            _add_photo(ws4, s.speaker_photo, col=3, row=ws4.max_row)

    _auto_width(ws4)
    ws4.column_dimensions["C"].width = 9

    # ============================================================
    # Sheet 5: 全体スケジュール (マトリクス忠実再現)
    # ============================================================
    ws5 = wb.create_sheet("全体スケジュール")
    room_list = sorted(rooms, key=lambda r: r.id)

    # --- フロントエンドと同じ列構成を構築 ---
    overall_sessions = [s for s in sessions if s.category == "overall"]
    session_only = [s for s in sessions if s.category in SESSION_CATS]
    # 動的カテゴリ別セッション
    cat_sessions_map = {ck: [s for s in sessions if s.category == ck] for ck in dynamic_cat_keys}
    all_schedule = overall_sessions + session_only
    for ck in dynamic_cat_keys:
        all_schedule += cat_sessions_map[ck]

    has_overall = len(overall_sessions) > 0
    # セッション部屋 (overall/動的カテゴリ 除外)
    sess_room_ids = dict.fromkeys(s.room_id for s in session_only if s.room)
    sess_rooms = [r for r in room_list if r.id in sess_room_ids]

    # 列定義: (type, label, fill, room_id_or_None)
    columns = []
    if has_overall:
        columns.append(("overall", "全体", HEADER_FILL_ORANGE, None))
    for r in sess_rooms:
        columns.append(("session", r.name, HEADER_FILL_INDIGO, r.id))
    # 動的カテゴリの部屋列
    for cat_obj in db_categories:
        ck = cat_obj.key
        cat_room_ids = dict.fromkeys(s.room_id for s in cat_sessions_map[ck] if s.room)
        cat_rooms = [r for r in room_list if r.id in cat_room_ids]
        fill = cat_fill_map[ck]
        for r in cat_rooms:
            columns.append((ck, f"{cat_obj.label}: {r.name}", fill, r.id))

    if not all_schedule:
        _auto_width(ws5)
    else:
        # --- 5分刻みの時間スロットを生成 ---
        from datetime import timedelta
        SLOT_MINUTES = 5
        slot_delta = timedelta(minutes=SLOT_MINUTES)

        all_starts = [s.start_time for s in all_schedule]
        all_ends = [s.end_time for s in all_schedule]
        min_time = min(all_starts)
        max_time = max(all_ends)
        # 5分単位に丸める
        min_time = min_time.replace(minute=(min_time.minute // SLOT_MINUTES) * SLOT_MINUTES, second=0, microsecond=0)
        if max_time.minute % SLOT_MINUTES:
            max_time = max_time.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1 if max_time.minute > 45 else 0, minutes=((max_time.minute // SLOT_MINUTES) + 1) * SLOT_MINUTES if max_time.minute % SLOT_MINUTES else 0)
        max_time = max_time.replace(second=0, microsecond=0)
        # もう少し安全に丸める
        import math
        max_minutes = max_time.hour * 60 + max_time.minute
        max_minutes = math.ceil(max_minutes / SLOT_MINUTES) * SLOT_MINUTES
        max_time = max_time.replace(hour=max_minutes // 60, minute=max_minutes % 60, second=0, microsecond=0)

        slots = []
        t = min_time
        while t < max_time:
            slots.append(t)
            t += slot_delta

        HEADER_ROW = 1
        DATA_START = 2  # データ開始行

        # --- ヘッダー行 ---
        ws5.cell(row=HEADER_ROW, column=1, value="時間")
        ws5.cell(row=HEADER_ROW, column=1).font = HEADER_FONT
        ws5.cell(row=HEADER_ROW, column=1).fill = HEADER_FILL_INDIGO
        ws5.cell(row=HEADER_ROW, column=1).alignment = CENTER
        ws5.cell(row=HEADER_ROW, column=1).border = THIN_BORDER

        for ci, (ctype, clabel, cfill, _) in enumerate(columns):
            cell = ws5.cell(row=HEADER_ROW, column=ci + 2, value=clabel)
            cell.font = HEADER_FONT
            cell.fill = cfill
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        # --- 時間ラベル (30分ごとに表示、15分スロットはセル結合) ---
        slot_row_map = {}  # slot_time -> excel_row
        for si, slot_t in enumerate(slots):
            excel_row = DATA_START + si
            slot_row_map[slot_t] = excel_row
            # 背景セルにボーダー
            for ci in range(len(columns)):
                cell = ws5.cell(row=excel_row, column=ci + 2, value="")
                cell.border = THIN_BORDER
            # 時間ラベル
            time_cell = ws5.cell(row=excel_row, column=1, value="")
            time_cell.border = THIN_BORDER
            time_cell.alignment = Alignment(horizontal="center", vertical="top")
            if slot_t.minute % 15 == 0:
                time_cell.value = _fmt(slot_t)
                time_cell.font = Font(bold=True, size=9)

        ws5.column_dimensions["A"].width = 8
        # 行高さ設定
        for si in range(len(slots)):
            ws5.row_dimensions[DATA_START + si].height = 14

        # --- カテゴリ別の色定義 ---
        CAT_FILL = {
            "overall": PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"),
            "general": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
            "tech": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
            "workshop": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
            "keynote": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
            "lt": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
            "panel": PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"),
        }
        CAT_FONT_COLOR = {
            "overall": "E65100",
            "general": "1A73E8", "tech": "1A73E8", "workshop": "1A73E8",
            "keynote": "1A73E8", "lt": "1A73E8", "panel": "1A73E8",
        }
        # 動的カテゴリの色を追加
        for cat_obj in db_categories:
            ck = cat_obj.key
            # 薄い背景色を生成（元の色を薄くする）
            hex_color = cat_obj.color.lstrip("#").upper()
            # セル背景: 薄い色
            r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
            lr = min(255, r + (255 - r) * 85 // 100)
            lg = min(255, g + (255 - g) * 85 // 100)
            lb = min(255, b + (255 - b) * 85 // 100)
            light_hex = f"{lr:02X}{lg:02X}{lb:02X}"
            CAT_FILL[ck] = PatternFill(start_color=light_hex, end_color=light_hex, fill_type="solid")
            CAT_FONT_COLOR[ck] = hex_color

        # --- セッションをマトリクスに配置 (セル結合) ---
        occupied_cells: set[tuple[int, int]] = set()  # (row, col) の使用済みセル

        for s in all_schedule:
            cat = s.category
            # 対応する列を見つける
            col_idx = None
            if cat == "overall":
                for ci, (ctype, _, _, _) in enumerate(columns):
                    if ctype == "overall":
                        col_idx = ci + 2
                        break
            elif cat in dynamic_cat_keys:
                for ci, (ctype, _, _, rid) in enumerate(columns):
                    if ctype == cat and rid == s.room_id:
                        col_idx = ci + 2
                        break
            else:  # session categories
                for ci, (ctype, _, _, rid) in enumerate(columns):
                    if ctype == "session" and rid == s.room_id:
                        col_idx = ci + 2
                        break

            if col_idx is None:
                continue

            # 行範囲を計算
            start_row = None
            end_row = None
            for slot_t, row in slot_row_map.items():
                if slot_t >= s.start_time and start_row is None:
                    start_row = row
                if slot_t < s.end_time:
                    end_row = row

            if start_row is None or end_row is None:
                # スロット範囲外 - 最も近いスロットを使う
                slot_times_list = sorted(slot_row_map.keys())
                if s.start_time <= slot_times_list[0]:
                    start_row = slot_row_map[slot_times_list[0]]
                else:
                    for st in slot_times_list:
                        if st <= s.start_time:
                            start_row = slot_row_map[st]
                if s.end_time >= slot_times_list[-1]:
                    end_row = slot_row_map[slot_times_list[-1]]
                else:
                    for st in slot_times_list:
                        if st < s.end_time:
                            end_row = slot_row_map[st]

            if start_row is None or end_row is None:
                continue

            # セル内容を構築
            staff_names = ", ".join(a.staff.name for a in s.assignments)
            time_str = f"{_fmt(s.start_time)}-{_fmt(s.end_time)}"
            if cat == "overall":
                content = f"{s.title}\n{time_str}"
                if s.required_staff == -1:
                    content += "\n【全員】"
                elif staff_names:
                    content += f"\n【{staff_names}】"
                if s.notes:
                    content += f"\n{s.notes}"
            else:
                content = f"{s.title}\n{time_str}"
                if s.speaker:
                    content += f"\n{s.speaker}"
                if staff_names:
                    content += f"\n【{staff_names}】"
                if not s.assignments and s.required_staff > 0:
                    content += "\n※未配置"

            # セルが既に使用済みならスキップ
            if (start_row, col_idx) in occupied_cells:
                continue
            for r in range(start_row, end_row + 1):
                occupied_cells.add((r, col_idx))

            # セル結合と書式設定
            if end_row > start_row:
                ws5.merge_cells(
                    start_row=start_row, start_column=col_idx,
                    end_row=end_row, end_column=col_idx,
                )

            cell = ws5.cell(row=start_row, column=col_idx, value=content)
            fill = CAT_FILL.get(cat, CAT_FILL["general"])
            font_color = CAT_FONT_COLOR.get(cat, "333333")
            cell.fill = fill
            cell.font = Font(size=9, color=font_color)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
            cell.border = THIN_BORDER

            # 結合セルの右下ボーダーも設定
            if end_row > start_row:
                for r in range(start_row, end_row + 1):
                    border_cell = ws5.cell(row=r, column=col_idx)
                    border_cell.border = THIN_BORDER

        # 列幅設定
        for ci, (ctype, clabel, _, _) in enumerate(columns):
            col_letter = get_column_letter(ci + 2)
            label_len = sum(2 if ord(c) > 127 else 1 for c in clabel)
            ws5.column_dimensions[col_letter].width = max(label_len + 2, 18)

    # ============================================================
    # 動的カテゴリ別シート
    # ============================================================
    for cat_obj in db_categories:
        ws_cat = wb.create_sheet(cat_obj.label)
        ws_cat.append(["タイトル", "時間", "場所", "英語", "必要人数", "配置人数", "担当スタッフ", "備考"])
        _apply_header(ws_cat, 1, cat_fill_map[cat_obj.key])

        for s in sessions:
            if s.category != cat_obj.key:
                continue
            staff_names = ", ".join(a.staff.name for a in s.assignments)
            assigned_count = len(s.assignments)
            status = "○" if assigned_count >= s.required_staff else f"不足({assigned_count}/{s.required_staff})"
            ws_cat.append([
                s.title,
                f"{_fmt(s.start_time)}-{_fmt(s.end_time)}",
                s.room.name if s.room else "",
                "○" if s.english_required else "",
                s.required_staff,
                status,
                staff_names,
                s.notes,
            ])
            _apply_border(ws_cat, ws_cat.max_row)

        _auto_width(ws_cat)

    # Excelファイルをバイトストリームに書き出し
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    ts = app_now().strftime("%Y%m%d_%H%M")
    filename = f"event_schedule_{ts}.xlsx"

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

